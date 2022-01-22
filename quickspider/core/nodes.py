from quickspider.core.Node import BaseNode, IterableNode
from scrapy import Selector
from requests import Session
from termcolor import colored
import openpyxl
import csv
import json
import os


#-------------> 网络节点 <--------------
class SessionNode(BaseNode):
    def __init__(self, _name, _session=None):
        super().__init__(_name)
        if not _session:
            _session = Session()
            _session.headers = {"User-Agent": "Google-Bot"}
        self._session = _session


class GetNode(SessionNode):
    def process_input(self, _input_url):
        try:
            resp = self._session.get(_input_url)
        except Exception as e:
            print(colored(str(e), "red"))
            resp = None
        return resp


class PostNode(SessionNode):
    def _set_data(self, _data: dict=None):
        if not _data:
            _data = {}
        self._data = _data

    def process_input(self, _input_url):
        try:
            resp = self._session.post(_input_url, json=self._data)
        except Exception as e:
            print(colored(str(e), "red"))
            resp = None
        return resp


#-------------> 解析节点 <--------------
class ParserNode(BaseNode):
    def __init__(self, _name, _mode):
        super().__init__(_name)
        self._set_mode(_mode)

    def _set_mode(self, _mode="dom"):
        self._mode = _mode

    def process_input(self, _response):
        if self._mode == "dom":
            return Selector(text=_response.text)
        if self._mode == "json":
            return _response.json()


class ParserDomNode(BaseNode):
    def __init__(self, _name, _mode="css", _parser=""):
        super().__init__(_name)
        self._set_mode(_mode)
        self._set_parser(_parser)

    def _set_mode(self, _mode="css"):
        self._mode = _mode

    def _set_parser(self, _parser):
        self._parser = _parser

    def process_input(self, _response):
        selector = _response
        if self._mode == "css":
            result = selector.css(self._parser)
        elif self._mode == "xpath":
            result = selector.xpath(self._parser)
        return result


class ParserJsonNode(BaseNode):
    pass


#-------------> 组合节点 <--------------
class ExtractNode(IterableNode):
    def process_input(self, _input_parser):
        for _sub_parser in _input_parser:
            yield _sub_parser


class SplitNode(IterableNode):
    def process_input(self, _input_parser):
        result = _input_parser.getall()
        return (i for i in result)


class ConcatNode(BaseNode):
    def process_input(self, _input_parser):
        result = "".join(_input_parser.getall())
        return result


#-------------> 输出测试节点 <--------------
class PrintNode(BaseNode):
    def process_input(self, _input):
        print(colored(_input, "green"))
        print("")
        return _input


class PageNode(IterableNode):
    def __init__(self, _name, _input, _start, _stop, _step=1):
        super().__init__(_name)
        self._set_input(_input)
        self._info = (_start, _stop, _step)

    def process_input(self, _input):
        _start, _stop, _step = self._info
        return (_input.format(page=p) for p in range(_start, _stop, _step))


#-------------> 读写节点 <--------------
FileReaderNode = IterableNode

class CsvReaderNode(FileReaderNode):
    def __init__(self, _name, _file, _column):
        super().__init__(_name)
        self._set_input(_file)
        self._column = _column

    def process_input(self, _input):
        if not os.path.exists(_input):
            raise Exception("Not Exist")
        with open(_input, mode="rt") as csvfile:
            csvdictreader = csv.DictReader(csvfile)
            for row in csvdictreader:
                yield row[self._column]
        

class ExcelReaderNode(FileReaderNode):
    def __init__(self, _name, _file, _column):
        super().__init__(_name)
        self._set_input(_file)
        self._column = _column

    def process_input(self, _input):
        wb = openpyxl.load_workbook(_input)
        # select active sheet by default
        sheet = wb.active
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(1, col).value == self._column:
                col_index = col
                break
        for row in range(2, sheet.max_rol + 1):
            cell = sheet.cell(row=row, column=col_indx)
            yield cell.value


class LineReaderNode(FileReaderNode):
    def __init__(self, _name, _file):
        super().__init__(_name)
        self._set_input(_file)

    def process_input(self, _input):
        with open(_input, "rt", encoding="utf8") as f:
            for line in f:
                yield line


class JsonWriterNode(BaseNode):
    def __init__(self, _name, _file):
        super().__init__(_name)
        self._file = _file
        self._container = []

    def init(self):
        if os.path.exists(self._file):
            os.remove(self._file)

    def process_input(self, _input):
        self._container.append(_input)
        return _input

    def post(self):
        with open(self._file, "wt", encoding="utf8") as f:
            json.dump(self._container, f, ensure_ascii=False, indent=4)

class CsvWriterNode(BaseNode):
    def __init__(self, _name, _file):
        super().__init__(_name)
        self._file = _file

    def init(self):
        if os.path.exists(self._file):
            os.remove(self._file)
        self._file_handel = open(self._file, "wt", encoding="utf8")
        self._csv_writer = csv.writer(self._file_handel)
        self._csv_writer.writerow(["result"])

    def process_input(self, _input):
        self._csv_writer.writerow([_input])
        return _input

    def post(self):
        self._file_handel.close()

# add your own node 

